from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

AFTERSALE_NO_COLUMN = "售后单号"
FINISHED_AT_COLUMN = "售后完结时间"
REQUIRED_COLUMNS = (AFTERSALE_NO_COLUMN, FINISHED_AT_COLUMN)
DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y/%m/%d",
)


@dataclass(frozen=True)
class MergeSummary:
    destination: Path
    total_rows: int
    unique_rows: int
    duplicate_rows: int
    daily_counts: dict[str, int]


def merge_tabular_exports(
    files: list[Path],
    destination: Path,
    *,
    timezone_name: str = "Asia/Shanghai",
) -> MergeSummary:
    total_rows = 0
    columns: list[str] = []
    merged_rows: list[dict[str, Any]] = []
    seen_aftersale_nos: set[str] = set()
    daily_counts: dict[str, int] = {}

    for file_path in files:
        file_columns, file_rows = _read_rows(file_path)
        _validate_required_columns(file_columns, file_path)

        for column in file_columns:
            if column not in columns:
                columns.append(column)

        for row_number, row in enumerate(file_rows, start=2):
            total_rows += 1
            aftersale_no = _require_text_value(
                row.get(AFTERSALE_NO_COLUMN),
                column_name=AFTERSALE_NO_COLUMN,
                file_path=file_path,
                row_number=row_number,
            )
            finished_at = _require_datetime_value(
                row.get(FINISHED_AT_COLUMN),
                column_name=FINISHED_AT_COLUMN,
                file_path=file_path,
                row_number=row_number,
                timezone_name=timezone_name,
            )
            if aftersale_no in seen_aftersale_nos:
                continue

            seen_aftersale_nos.add(aftersale_no)
            merged_rows.append(row)
            day = finished_at.date().isoformat()
            daily_counts[day] = daily_counts.get(day, 0) + 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(columns)
    for row in merged_rows:
        sheet.append([row.get(column) for column in columns])
    workbook.save(destination)

    return MergeSummary(
        destination=destination,
        total_rows=total_rows,
        unique_rows=len(merged_rows),
        duplicate_rows=total_rows - len(merged_rows),
        daily_counts=daily_counts,
    )


def _read_rows(file_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if file_path.suffix.lower() == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fieldnames = [str(value) if value is not None else "" for value in (reader.fieldnames or [])]
            return fieldnames, list(reader)

    if file_path.suffix.lower() == ".xlsx":
        workbook = load_workbook(file_path)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        headers = [str(value) if value is not None else "" for value in values[0]]
        rows: list[dict[str, Any]] = []
        for value_row in values[1:]:
            row: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                row[header] = value_row[idx] if idx < len(value_row) else None
            rows.append(row)
        return headers, rows

    raise ValueError(f"Unsupported file type for merge: {file_path.suffix}")


def _validate_required_columns(columns: list[str], file_path: Path) -> None:
    for column in REQUIRED_COLUMNS:
        if column not in columns:
            raise ValueError(f"{file_path.name}: missing required column: {column}")


def _require_text_value(
    value: Any,
    *,
    column_name: str,
    file_path: Path,
    row_number: int,
) -> str:
    if value is None:
        raise ValueError(f"{file_path.name} row {row_number}: blank value for required column {column_name}")

    text = str(value).strip()
    if not text:
        raise ValueError(f"{file_path.name} row {row_number}: blank value for required column {column_name}")
    return text


def _require_datetime_value(
    value: Any,
    *,
    column_name: str,
    file_path: Path,
    row_number: int,
    timezone_name: str,
) -> datetime:
    if value is None:
        raise ValueError(f"{file_path.name} row {row_number}: blank value for required column {column_name}")

    timezone = ZoneInfo(timezone_name)
    parsed = _parse_datetime_value(value)
    if parsed is None:
        raise ValueError(
            f"{file_path.name} row {row_number}: invalid datetime for required column {column_name}: {value!r}"
        )
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone)
    return parsed.astimezone(timezone)


def _parse_datetime_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        try:
            return from_excel(value)
        except (TypeError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            pass
        for fmt in DATETIME_FORMATS:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None
