from __future__ import annotations

from contextlib import redirect_stdout
from datetime import datetime
import io
import json
from pathlib import Path
import tempfile
import unittest
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

from aftersale_exporter.job import MANIFEST_RUN_SEPARATOR_PREFIX, AftersaleExportJob
from aftersale_exporter.workflow import OverLimitError, TaskPollResult


LOCAL_TZ = ZoneInfo("Asia/Shanghai")


def local_ts(value: str) -> int:
    return int(datetime.strptime(value, "%Y-%m-%d %H:%M:%S").replace(tzinfo=LOCAL_TZ).timestamp())


class FakeClock:
    def __init__(self) -> None:
        self.now = 0.0

    def monotonic(self) -> float:
        return self.now

    def sleep(self, seconds: float) -> None:
        self.now += seconds


class MixedFormatFakeService:
    def __init__(self) -> None:
        self.submissions: list[tuple[int, int]] = []
        self.count_requests: list[tuple[int, int]] = []

    def create_export(self, start_ts: int, end_ts: int) -> str:
        self.submissions.append((start_ts, end_ts))
        if (start_ts, end_ts) == (0, 3):
            raise OverLimitError("too many rows")
        return f"task-{start_ts}-{end_ts}"

    def wait_for_task(self, task_id: str, poll_interval: float, timeout: float, status_callback=None):
        raise AssertionError("job workflow should poll tasks incrementally")

    def poll_task(self, task_id: str) -> TaskPollResult:
        if task_id == "task-0-1":
            return TaskPollResult(
                requested_at_ts=1234567890,
                result_text="文件已生成",
                is_complete=True,
                download_name="left.csv",
            )
        return TaskPollResult(
            requested_at_ts=1234567891,
            result_text="文件已生成",
            is_complete=True,
            download_name="right.xlsx",
        )

    def download_export(self, task_id: str, destination: Path) -> Path:
        if destination.suffix == ".csv":
            destination.write_text(
                "\n".join(
                    [
                        "售后单号,售后完结时间,left_only",
                        "A1,1970-01-01 08:00:00,L1",
                        "A2,1970-01-01 08:00:03,L2",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )
            return destination

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["售后单号", "售后完结时间", "right_only"])
        sheet.append(["A2", datetime(1970, 1, 1, 8, 0, 3), "R2"])
        sheet.append(["A3", datetime(1970, 1, 1, 8, 0, 5), "R3"])
        workbook.save(destination)
        return destination

    def count_aftersales(self, start_ts: int, end_ts: int) -> int:
        self.count_requests.append((start_ts, end_ts))
        return len(self.count_requests) + 1


class UnsupportedFormatFakeService:
    def __init__(self) -> None:
        self.count_requests: list[tuple[int, int]] = []

    def create_export(self, start_ts: int, end_ts: int) -> str:
        return f"task-{start_ts}-{end_ts}"

    def wait_for_task(self, task_id: str, poll_interval: float, timeout: float, status_callback=None):
        raise AssertionError("job workflow should poll tasks incrementally")

    def poll_task(self, task_id: str) -> TaskPollResult:
        return TaskPollResult(
            requested_at_ts=1234567890,
            result_text="文件已生成",
            is_complete=True,
            download_name="bad.txt",
        )

    def download_export(self, task_id: str, destination: Path) -> Path:
        destination.write_text("not-a-tabular-export", encoding="utf-8")
        return destination

    def count_aftersales(self, start_ts: int, end_ts: int) -> int:
        self.count_requests.append((start_ts, end_ts))
        return 9


class PartialFailureFakeService:
    def __init__(self) -> None:
        self.calls: list[str] = []
        self.count_requests: list[tuple[int, int]] = []

    def create_export(self, start_ts: int, end_ts: int) -> str:
        self.calls.append(f"create:{start_ts}-{end_ts}")
        if (start_ts, end_ts) == (0, 3):
            raise OverLimitError("too many rows")
        return f"task-{start_ts}-{end_ts}"

    def wait_for_task(self, task_id: str, poll_interval: float, timeout: float, status_callback=None):
        raise AssertionError("job workflow should poll tasks incrementally")

    def poll_task(self, task_id: str) -> TaskPollResult:
        return TaskPollResult(
            requested_at_ts=1234567890,
            result_text="文件已生成",
            is_complete=True,
            download_name=f"{task_id}.csv",
        )

    def download_export(self, task_id: str, destination: Path) -> Path:
        if task_id == "task-2-3":
            raise RuntimeError("disk full")
        destination.write_text(
            "\n".join(
                [
                    "售后单号,售后完结时间,value",
                    "A1,1970-01-01 08:00:00,ok",
                ]
            )
            + "\n",
            encoding="utf-8",
        )
        return destination

    def count_aftersales(self, start_ts: int, end_ts: int) -> int:
        self.count_requests.append((start_ts, end_ts))
        return 5


class DailyCountFakeService(MixedFormatFakeService):
    def __init__(
        self,
        *,
        count_totals: dict[tuple[int, int], int] | None = None,
        count_failures: dict[tuple[int, int], Exception] | None = None,
        export_rows: dict[str, list[tuple[str, datetime | str, str]]] | None = None,
    ) -> None:
        super().__init__()
        self.count_totals = count_totals or {}
        self.count_failures = count_failures or {}
        self.export_rows = export_rows or {
            "task-default": [("A1", "1970-01-01 08:00:00", "ok")],
        }

    def create_export(self, start_ts: int, end_ts: int) -> str:
        return f"task-{start_ts}-{end_ts}"

    def poll_task(self, task_id: str) -> TaskPollResult:
        return TaskPollResult(
            requested_at_ts=1234567890,
            result_text="文件已生成",
            is_complete=True,
            download_name=f"{task_id}.csv",
        )

    def download_export(self, task_id: str, destination: Path) -> Path:
        rows = self.export_rows.get(task_id)
        if rows is None:
            rows = next(iter(self.export_rows.values()))
        lines = ["售后单号,售后完结时间,value"]
        for order_no, finished_at, value in rows:
            cell = finished_at if isinstance(finished_at, str) else finished_at.strftime("%Y-%m-%d %H:%M:%S")
            lines.append(f"{order_no},{cell},{value}")
        destination.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return destination

    def count_aftersales(self, start_ts: int, end_ts: int) -> int:
        key = (start_ts, end_ts)
        self.count_requests.append(key)
        if key in self.count_failures:
            raise self.count_failures[key]
        return self.count_totals[key]


class RemediationFakeService:
    """Fake service that returns different export data on second call for same range."""

    def __init__(
        self,
        *,
        count_totals: dict[tuple[int, int], int],
        first_export_rows: dict[str, list[tuple[str, str, str]]],
        second_export_rows: dict[str, list[tuple[str, str, str]]] | None = None,
    ) -> None:
        self.count_totals = count_totals
        self.count_requests: list[tuple[int, int]] = []
        self.submissions: list[tuple[int, int]] = []
        self.first_export_rows = first_export_rows
        self.second_export_rows = second_export_rows or {}
        self._export_call_counts: dict[tuple[int, int], int] = {}

    def create_export(self, start_ts: int, end_ts: int) -> str:
        self.submissions.append((start_ts, end_ts))
        return f"task-{start_ts}-{end_ts}"

    def wait_for_task(self, task_id, poll_interval, timeout, status_callback=None):
        raise AssertionError("job workflow should poll tasks incrementally")

    def poll_task(self, task_id: str) -> TaskPollResult:
        return TaskPollResult(
            requested_at_ts=1234567890,
            result_text="文件已生成",
            is_complete=True,
            download_name=f"{task_id}.csv",
        )

    def download_export(self, task_id: str, destination: Path) -> Path:
        key = tuple(int(x) for x in task_id.replace("task-", "").split("-"))
        call_index = self._export_call_counts.get(key, 0)
        self._export_call_counts[key] = call_index + 1

        rows_map = self.first_export_rows if call_index == 0 else self.second_export_rows
        rows = rows_map.get(task_id)
        if rows is None:
            rows = next(iter(rows_map.values()))
        lines = ["售后单号,售后完结时间,value"]
        for order_no, finished_at, value in rows:
            lines.append(f"{order_no},{finished_at},{value}")
        destination.write_text("\n".join(lines) + "\n", encoding="utf-8")
        return destination

    def count_aftersales(self, start_ts: int, end_ts: int) -> int:
        key = (start_ts, end_ts)
        self.count_requests.append(key)
        return self.count_totals[key]


def load_latest_manifest_snapshot(manifest_path: Path) -> dict[str, object]:
    manifest_text = manifest_path.read_text(encoding="utf-8")
    marker_index = manifest_text.rfind(MANIFEST_RUN_SEPARATOR_PREFIX)
    if marker_index == -1:
        return json.loads(manifest_text)
    latest_start = manifest_text.find("\n", marker_index)
    return json.loads(manifest_text[latest_start + 1 :].strip())


class AftersaleExportJobTests(unittest.TestCase):
    def test_job_writes_manifest_and_deduplicated_merged_xlsx_for_split_exports(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            job = AftersaleExportJob(
                service=MixedFormatFakeService(),
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                result = job.run(start_ts=0, end_ts=3)

            manifest_path = Path(tmpdir) / "manifest.json"
            merged_path = Path(tmpdir) / "merged.xlsx"
            raw_dir = Path(tmpdir) / "raw"

            self.assertTrue(manifest_path.exists())
            self.assertTrue(merged_path.exists())
            self.assertTrue((raw_dir / "left.csv").exists())
            self.assertTrue((raw_dir / "right.xlsx").exists())
            self.assertEqual(result.segment_count, 2)

            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["summary"]["segment_count"], 2)
            self.assertEqual(manifest["summary"]["failed_count"], 0)
            self.assertEqual(manifest["summary"]["daily_count_days"], 1)
            self.assertEqual(manifest["summary"]["daily_count_failed_days"], 0)
            self.assertEqual(
                [(item["start_ts"], item["end_ts"]) for item in manifest["segments"]],
                [(0, 1), (2, 3)],
            )
            self.assertEqual(
                manifest["daily_counts"],
                [
                    {
                        "date": "1970-01-01",
                        "start_ts": 0,
                        "end_ts": 3,
                        "status": "counted",
                        "total": 2,
                    }
                ],
            )

            workbook = load_workbook(merged_path)
            rows = list(workbook.active.iter_rows(values_only=True))
            self.assertEqual(rows[0], ("售后单号", "售后完结时间", "left_only", "right_only"))
            self.assertEqual(len(rows), 4)
            self.assertEqual(rows[1][0], "A1")
            self.assertEqual(rows[2][0], "A2")
            self.assertEqual(rows[2][2], "L2")
            self.assertIsNone(rows[2][3])
            self.assertEqual(rows[3][0], "A3")

            output = stdout.getvalue()
            self.assertIn("1970-01-01 | manifest=2 | merged=3 | MISMATCH", output)
            self.assertIn("去重汇总 | total=4 | unique=3 | duplicates=1", output)
            self.assertIn("比对汇总 | match=0 | mismatch=1 | skipped=0", output)

    def test_job_keeps_raw_files_and_records_merge_error(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            job = AftersaleExportJob(
                service=UnsupportedFormatFakeService(),
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
            )

            result = job.run(start_ts=10, end_ts=10)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )

            self.assertEqual(result.segment_count, 1)
            self.assertTrue((Path(tmpdir) / "raw" / "bad.txt").exists())
            self.assertFalse((Path(tmpdir) / "merged.xlsx").exists())
            self.assertIn("merge_error", manifest["summary"])
            self.assertIn("Unsupported file type", manifest["summary"]["merge_error"])
            self.assertEqual(manifest["summary"]["daily_count_days"], 1)
            self.assertEqual(manifest["summary"]["daily_count_failed_days"], 0)
            self.assertEqual(manifest["daily_counts"][0]["total"], 9)

    def test_job_persists_manifest_before_raising_partial_failure(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            job = AftersaleExportJob(
                service=PartialFailureFakeService(),
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
            )

            with self.assertRaisesRegex(RuntimeError, "disk full"):
                job.run(start_ts=0, end_ts=3)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )

            self.assertEqual(manifest["summary"]["failed_count"], 1)
            self.assertEqual(
                [(item["start_ts"], item["end_ts"], item["status"]) for item in manifest["segments"]],
                [(0, 1, "downloaded"), (2, 3, "failed")],
            )
            self.assertEqual(manifest["failures"][0]["message"], "disk full")

    def test_job_emits_progress_events(self) -> None:
        events: list[str] = []
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            job = AftersaleExportJob(
                service=MixedFormatFakeService(),
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                progress_callback=lambda event_name, payload: events.append(event_name),
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
            )

            job.run(start_ts=0, end_ts=3)

        self.assertEqual(events[0], "split")
        self.assertIn("waiting_task", events)
        self.assertIn("task_polled", events)
        self.assertIn("waiting_export_gap", events)
        self.assertEqual(events.count("downloaded"), 4)
        self.assertIn("remediation", events)
        self.assertIn("counted", events)

    def test_job_records_daily_counts_across_multiple_local_days(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            first_start = local_ts("2026-05-01 10:00:00")
            first_end = local_ts("2026-05-01 23:59:59")
            second_start = local_ts("2026-05-02 00:00:00")
            second_end = local_ts("2026-05-02 23:59:59")
            third_start = local_ts("2026-05-03 00:00:00")
            third_end = local_ts("2026-05-03 12:00:00")
            service = DailyCountFakeService(
                count_totals={
                    (first_start, first_end): 11,
                    (second_start, second_end): 22,
                    (third_start, third_end): 33,
                },
                export_rows={
                    f"task-{first_start}-{third_end}": [
                        ("A1", "2026-05-01 12:00:00", "v1"),
                        ("A2", "2026-05-02 12:00:00", "v2"),
                        ("A3", "2026-05-03 10:00:00", "v3"),
                    ]
                },
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                job.run(start_ts=first_start, end_ts=third_end)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )

        self.assertEqual(
            service.count_requests,
            [
                (first_start, first_end),
                (second_start, second_end),
                (third_start, third_end),
            ],
        )
        self.assertEqual(manifest["summary"]["daily_count_days"], 3)
        self.assertEqual(manifest["summary"]["daily_count_failed_days"], 0)
        self.assertEqual(
            manifest["daily_counts"],
            [
                {
                    "date": "2026-05-01",
                    "start_ts": first_start,
                    "end_ts": first_end,
                    "status": "counted",
                    "total": 11,
                },
                {
                    "date": "2026-05-02",
                    "start_ts": second_start,
                    "end_ts": second_end,
                    "status": "counted",
                    "total": 22,
                },
                {
                    "date": "2026-05-03",
                    "start_ts": third_start,
                    "end_ts": third_end,
                    "status": "counted",
                    "total": 33,
                },
            ],
        )
        output = stdout.getvalue()
        self.assertIn("2026-05-01 | manifest=11 | merged=1 | MISMATCH", output)
        self.assertIn("2026-05-02 | manifest=22 | merged=1 | MISMATCH", output)
        self.assertIn("2026-05-03 | manifest=33 | merged=1 | MISMATCH", output)

    def test_job_prints_match_results_for_consistent_daily_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            start_ts = local_ts("2026-05-01 00:00:00")
            end_ts = local_ts("2026-05-01 23:59:59")
            service = DailyCountFakeService(
                count_totals={(start_ts, end_ts): 2},
                export_rows={
                    f"task-{start_ts}-{end_ts}": [
                        ("A1", "2026-05-01 09:00:00", "v1"),
                        ("A2", "2026-05-01 10:00:00", "v2"),
                    ]
                },
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                job.run(start_ts=start_ts, end_ts=end_ts)

        output = stdout.getvalue()
        self.assertIn("2026-05-01 | manifest=2 | merged=2 | MATCH", output)
        self.assertIn("比对汇总 | match=1 | mismatch=0 | skipped=0", output)

    def test_job_records_partial_daily_count_failures_without_failing_export(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            first_start = local_ts("2026-05-01 08:00:00")
            first_end = local_ts("2026-05-01 23:59:59")
            second_start = local_ts("2026-05-02 00:00:00")
            second_end = local_ts("2026-05-02 08:00:00")
            service = DailyCountFakeService(
                count_totals={(first_start, first_end): 7},
                count_failures={(second_start, second_end): RuntimeError("count failed")},
                export_rows={
                    f"task-{first_start}-{second_end}": [
                        ("A1", "2026-05-01 12:00:00", "v1"),
                        ("A2", "2026-05-02 02:00:00", "v2"),
                    ]
                },
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                result = job.run(start_ts=first_start, end_ts=second_end)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )

        self.assertEqual(result.segment_count, 1)
        self.assertEqual(manifest["summary"]["failed_count"], 0)
        self.assertEqual(manifest["summary"]["daily_count_days"], 2)
        self.assertEqual(manifest["summary"]["daily_count_failed_days"], 1)
        self.assertEqual(manifest["daily_counts"][0]["status"], "counted")
        self.assertEqual(manifest["daily_counts"][0]["total"], 7)
        self.assertEqual(manifest["daily_counts"][1]["status"], "failed")
        self.assertEqual(manifest["daily_counts"][1]["error_type"], "RuntimeError")
        self.assertEqual(manifest["daily_counts"][1]["message"], "count failed")
        self.assertIn("2026-05-02 | manifest=FAILED | merged=1 | SKIPPED", stdout.getvalue())

    def test_job_keeps_previous_manifest_run_with_separator(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            first_clock = FakeClock()
            first_job = AftersaleExportJob(
                service=DailyCountFakeService(
                    count_totals={(0, 0): 1},
                    export_rows={f"task-{0}-{0}": [("A1", "1970-01-01 08:00:00", "v1")]},
                ),
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=first_clock.sleep,
                time_fn=first_clock.monotonic,
            )
            second_clock = FakeClock()
            second_job = AftersaleExportJob(
                service=DailyCountFakeService(
                    count_totals={(1, 1): 1},
                    export_rows={f"task-{1}-{1}": [("A2", "1970-01-01 08:00:01", "v2")]},
                ),
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=second_clock.sleep,
                time_fn=second_clock.monotonic,
            )

            first_job.run(start_ts=0, end_ts=0)
            first_manifest_text = (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")

            second_job.run(start_ts=1, end_ts=1)

            manifest_path = Path(tmpdir) / "manifest.json"
            manifest_text = manifest_path.read_text(encoding="utf-8")
            latest_manifest = load_latest_manifest_snapshot(manifest_path)

        self.assertIn(first_manifest_text.strip(), manifest_text)
        self.assertIn(MANIFEST_RUN_SEPARATOR_PREFIX, manifest_text)
        self.assertEqual(latest_manifest["summary"]["segment_count"], 1)
        self.assertEqual(latest_manifest["segments"][0]["start_ts"], 1)

    def test_remediation_resolves_single_day_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            start_ts = local_ts("2026-05-01 00:00:00")
            end_ts = local_ts("2026-05-01 23:59:59")
            service = RemediationFakeService(
                count_totals={(start_ts, end_ts): 3},
                first_export_rows={
                    f"task-{start_ts}-{end_ts}": [
                        ("A1", "2026-05-01 09:00:00", "v1"),
                    ]
                },
                second_export_rows={
                    f"task-{start_ts}-{end_ts}": [
                        ("A2", "2026-05-01 10:00:00", "v2"),
                        ("A3", "2026-05-01 11:00:00", "v3"),
                    ]
                },
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                job.run(start_ts=start_ts, end_ts=end_ts)

            output = stdout.getvalue()
            self.assertIn(
                "2026-05-01 | manifest=3 | merged=1 | MISMATCH", output
            )
            self.assertIn("开始修复 1 天 mismatch", output)
            self.assertIn("RESOLVED", output)
            self.assertIn(
                "修复汇总 | 共 1 天 mismatch | resolved=1 | unresolved=0", output
            )
            self.assertNotIn("UNRESOLVED", output)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )
            self.assertTrue(manifest["summary"]["remediation"]["attempted"])
            self.assertEqual(
                manifest["summary"]["remediation"]["resolved_count"], 1
            )
            self.assertEqual(
                manifest["summary"]["remediation"]["unresolved_count"], 0
            )

    def test_remediation_unresolved_when_redownload_yields_same_data(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            start_ts = local_ts("2026-05-01 00:00:00")
            end_ts = local_ts("2026-05-01 23:59:59")
            rows = [("A1", "2026-05-01 09:00:00", "v1")]
            service = RemediationFakeService(
                count_totals={(start_ts, end_ts): 3},
                first_export_rows={f"task-{start_ts}-{end_ts}": rows},
                second_export_rows={f"task-{start_ts}-{end_ts}": rows},
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                job.run(start_ts=start_ts, end_ts=end_ts)

            output = stdout.getvalue()
            self.assertIn("UNRESOLVED", output)
            self.assertIn("resolved=0 | unresolved=1", output)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )
            self.assertTrue(manifest["summary"]["remediation"]["attempted"])
            self.assertEqual(
                manifest["summary"]["remediation"]["unresolved_count"], 1
            )

    def test_no_remediation_when_all_days_match(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            start_ts = local_ts("2026-05-01 00:00:00")
            end_ts = local_ts("2026-05-01 23:59:59")
            service = DailyCountFakeService(
                count_totals={(start_ts, end_ts): 2},
                export_rows={
                    f"task-{start_ts}-{end_ts}": [
                        ("A1", "2026-05-01 09:00:00", "v1"),
                        ("A2", "2026-05-01 10:00:00", "v2"),
                    ]
                },
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                job.run(start_ts=start_ts, end_ts=end_ts)

            output = stdout.getvalue()
            self.assertIn("MATCH", output)
            self.assertNotIn("开始修复", output)

            manifest = json.loads(
                (Path(tmpdir) / "manifest.json").read_text(encoding="utf-8")
            )
            self.assertFalse(manifest["summary"]["remediation"]["attempted"])

    def test_skipped_days_not_in_remediation(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            clock = FakeClock()
            first_start = local_ts("2026-05-01 08:00:00")
            first_end = local_ts("2026-05-01 23:59:59")
            second_start = local_ts("2026-05-02 00:00:00")
            second_end = local_ts("2026-05-02 08:00:00")
            service = DailyCountFakeService(
                count_totals={(first_start, first_end): 1},
                count_failures={
                    (second_start, second_end): RuntimeError("boom")
                },
                export_rows={
                    f"task-{first_start}-{second_end}": [
                        ("A1", "2026-05-01 12:00:00", "v1"),
                        ("A2", "2026-05-02 02:00:00", "v2"),
                    ]
                },
            )
            job = AftersaleExportJob(
                service=service,
                out_dir=Path(tmpdir),
                poll_interval=0.01,
                task_timeout=1.0,
                sleep_fn=clock.sleep,
                time_fn=clock.monotonic,
                timezone_name="Asia/Shanghai",
            )

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                job.run(start_ts=first_start, end_ts=second_end)

            output = stdout.getvalue()
            self.assertIn("SKIPPED", output)
            self.assertNotIn("开始修复", output)


if __name__ == "__main__":
    unittest.main()
