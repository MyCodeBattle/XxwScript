from __future__ import annotations

from datetime import datetime
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from aftersale_exporter.merge import MergeSummary, merge_tabular_exports


class MergeTests(unittest.TestCase):
    def test_merge_deduplicates_by_aftersale_no_and_collects_daily_counts(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "left.csv"
            csv_path.write_text(
                "\n".join(
                    [
                        "售后单号,售后完结时间,来源",
                        "A1,2026-05-01 09:00:00,left",
                        "A2,2026-05-02 10:00:00,left",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            xlsx_path = Path(tmpdir) / "right.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["售后单号", "售后完结时间", "来源"])
            sheet.append(["A1", datetime(2026, 5, 1, 11, 30, 0), "right"])
            sheet.append(["A3", datetime(2026, 5, 2, 15, 0, 0), "right"])
            workbook.save(xlsx_path)

            destination = Path(tmpdir) / "merged.xlsx"
            summary = merge_tabular_exports([csv_path, xlsx_path], destination)

            self.assertIsInstance(summary, MergeSummary)
            self.assertEqual(summary.total_rows, 4)
            self.assertEqual(summary.unique_rows, 3)
            self.assertEqual(summary.duplicate_rows, 1)
            self.assertEqual(
                summary.daily_counts,
                {
                    "2026-05-01": 1,
                    "2026-05-02": 2,
                },
            )

            merged_rows = list(load_workbook(destination).active.iter_rows(values_only=True))
            self.assertEqual(merged_rows[0], ("售后单号", "售后完结时间", "来源"))
            self.assertEqual(merged_rows[1][0], "A1")
            self.assertEqual(merged_rows[1][2], "left")
            self.assertEqual(merged_rows[2][0], "A2")
            self.assertEqual(merged_rows[3][0], "A3")

    def test_merge_rejects_missing_required_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "missing.csv"
            csv_path.write_text(
                "\n".join(
                    [
                        "售后单号,来源",
                        "A1,left",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, "missing required column: 售后完结时间"):
                merge_tabular_exports([csv_path], Path(tmpdir) / "merged.xlsx")

    def test_merge_rejects_blank_required_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "blank.csv"
            csv_path.write_text(
                "\n".join(
                    [
                        "售后单号,售后完结时间",
                        ",2026-05-01 09:00:00",
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ValueError, "blank value for required column 售后单号"):
                merge_tabular_exports([csv_path], Path(tmpdir) / "merged.xlsx")


if __name__ == "__main__":
    unittest.main()
